from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from stt_desktop.backups import BackupService
from stt_desktop.storage.project import ProjectError, ProjectRepository, ProjectWorkspace, utc_now, uuid7


MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_IMPORT_ROWS = 100_000
MAX_IMPORT_COLUMNS = 50
MAX_CELL_LENGTH = 10_000


@dataclass(frozen=True)
class ImportField:
    name: str
    aliases: tuple[str, ...]
    required: bool = False
    converter: Callable[[Any], Any] | None = None


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: Any) -> str | None:
    result = _text(value)
    return result or None


def _integer(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    number = float(text)
    if not math.isfinite(number) or not number.is_integer():
        raise ValueError("不是有限整数")
    return int(number)


def _positive_integer(value: Any) -> int:
    result = _integer(value)
    if result is None or result <= 0:
        raise ValueError("必须为正整数")
    return result


def _nonnegative_integer(value: Any) -> int:
    result = _integer(value)
    if result is None or result < 0:
        raise ValueError("必须为非负整数")
    return result


def _bits(value: Any, default: str, maximum: int) -> str:
    result = _text(value) or default
    if len(result) > maximum or any(bit not in {"0", "1"} for bit in result):
        raise ValueError(f"必须是最多 {maximum} 位的 0/1 字符串")
    return result


def _boolean(value: Any) -> int:
    text = _text(value).lower()
    if text in {"1", "true", "yes", "y", "是", "需要"}:
        return 1
    if text in {"0", "false", "no", "n", "否", "不需要", ""}:
        return 0
    raise ValueError("必须为是/否或 true/false")


def _status(value: Any) -> str:
    text = _text(value).lower()
    aliases = {"": "active", "启用": "active", "在职": "active", "active": "active", "停用": "inactive", "离职": "inactive", "inactive": "inactive"}
    if text not in aliases:
        raise ValueError("状态必须为启用/停用或 active/inactive")
    return aliases[text]


IMPORT_SCHEMAS: dict[str, tuple[ImportField, ...]] = {
    "teacher": (
        ImportField("employee_no", ("工号", "教师工号", "employee_no", "employee no"), converter=_optional_text),
        ImportField("name", ("姓名", "教师姓名", "name"), required=True, converter=_text),
        ImportField("department", ("部门", "院系", "department"), converter=_optional_text),
        ImportField("status", ("状态", "status"), converter=_status),
    ),
    "subject": (
        ImportField("name", ("科目名称", "课程名称", "名称", "name"), required=True, converter=_text),
        ImportField("code", ("科目代码", "课程代码", "代码", "code"), converter=_optional_text),
        ImportField("category", ("分类", "类别", "category"), converter=lambda value: _text(value) or "general"),
        ImportField("default_duration_slots", ("默认连续课时", "连续课时", "default_duration_slots"), converter=lambda value: _positive_integer(value) if _text(value) else 1),
        ImportField("requires_special_room", ("需要专用教室", "专用教室", "requires_special_room"), converter=_boolean),
    ),
    "grade": (
        ImportField("name", ("年级名称", "名称", "name"), required=True, converter=_text),
        ImportField("code", ("年级代码", "代码", "code"), converter=_optional_text),
        ImportField("sort_order", ("排序", "sort_order"), converter=lambda value: _integer(value) or 0),
    ),
    "room_type": (
        ImportField("name", ("教室类型名称", "类型名称", "名称", "name"), required=True, converter=_text),
        ImportField("code", ("类型代码", "代码", "code"), converter=_optional_text),
        ImportField("description", ("说明", "描述", "description"), converter=_optional_text),
    ),
    "room": (
        ImportField("name", ("教室名称", "名称", "name"), required=True, converter=_text),
        ImportField("room_no", ("教室编号", "门牌号", "room_no"), converter=_optional_text),
        ImportField("capacity", ("容量", "capacity"), converter=_integer),
        ImportField("status", ("状态", "status"), converter=_status),
        ImportField("room_type_name", ("教室类型", "类型", "room_type_name"), converter=_optional_text),
    ),
    "homeroom": (
        ImportField("name", ("班级名称", "名称", "name"), required=True, converter=_text),
        ImportField("group_name", ("分组", "组别", "group_name"), converter=_optional_text),
        ImportField("student_count", ("学生人数", "人数", "student_count"), converter=_integer),
        ImportField("status", ("状态", "status"), converter=_status),
        ImportField("grade_name", ("年级", "年级名称", "grade_name"), converter=_optional_text),
        ImportField("term_name", ("学期", "学期名称", "term_name"), converter=_optional_text),
        ImportField("head_teacher_name", ("班主任", "班主任姓名", "head_teacher_name"), converter=_optional_text),
        ImportField("default_room_name", ("默认教室", "教室", "default_room_name"), converter=_optional_text),
    ),
    "course_plan": (
        ImportField("term_name", ("学期", "学期名称", "term_name"), converter=_optional_text),
        ImportField("homeroom_name", ("班级", "班级名称", "homeroom_name"), required=True, converter=_text),
        ImportField("subject_name", ("科目", "科目名称", "subject_name"), required=True, converter=_text),
        ImportField("weekly_slots", ("每周课时", "周课时", "weekly_slots"), required=True, converter=_nonnegative_integer),
        ImportField("duration_slots", ("连续课时", "每次课时", "duration_slots"), converter=lambda value: _positive_integer(value) if _text(value) else 1),
        ImportField("allow_double_period", ("允许连堂", "连堂", "allow_double_period"), converter=_boolean),
        ImportField("priority", ("优先级", "priority"), converter=lambda value: _integer(value) or 0),
        ImportField("week_bits", ("周次位图", "week_bits"), converter=lambda value: _bits(value, "1" * 20, 60)),
        ImportField("day_bits", ("星期位图", "day_bits"), converter=lambda value: _bits(value, "11111", 7)),
    ),
    "teaching_task": (
        ImportField("term_name", ("学期", "学期名称", "term_name"), converter=_optional_text),
        ImportField("homeroom_name", ("班级", "班级名称", "homeroom_name"), required=True, converter=_text),
        ImportField("subject_name", ("科目", "科目名称", "subject_name"), required=True, converter=_text),
        ImportField("primary_teacher_name", ("主讲教师", "教师", "教师姓名", "primary_teacher_name"), required=True, converter=_text),
        ImportField("weekly_slots", ("每周课时", "周课时", "weekly_slots"), required=True, converter=_nonnegative_integer),
        ImportField("duration_slots", ("连续课时", "每次课时", "duration_slots"), converter=lambda value: _positive_integer(value) if _text(value) else 1),
        ImportField("required_room_type_name", ("要求教室类型", "教室类型", "required_room_type_name"), converter=_optional_text),
        ImportField("fixed_room_name", ("固定教室", "教室", "fixed_room_name"), converter=_optional_text),
        ImportField("status", ("状态", "status"), converter=_status),
        ImportField("week_bits", ("周次位图", "week_bits"), converter=lambda value: _bits(value, "1" * 20, 60)),
        ImportField("day_bits", ("星期位图", "day_bits"), converter=lambda value: _bits(value, "11111", 7)),
    ),
}


class ImportService:
    def __init__(self, project: ProjectRepository, workspace: ProjectWorkspace) -> None:
        self.project = project
        self.workspace = workspace

    def preview_file(
        self,
        *,
        source_path: str,
        entity_type: str,
        mapping: dict[str, str] | None = None,
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        if entity_type not in IMPORT_SCHEMAS:
            raise ProjectError("导入类型不受支持")
        source = Path(source_path).expanduser().resolve()
        if not source.is_file() or source.suffix.lower() not in {".csv", ".xlsx"}:
            raise ProjectError("请选择有效的 CSV 或 XLSX 文件")
        size = source.stat().st_size
        if size <= 0 or size > MAX_IMPORT_BYTES:
            raise ProjectError("导入文件必须大于 0 且不超过 20 MiB")
        job_id = uuid7()
        relative = f"attachments/imports/{job_id}{source.suffix.lower()}"
        stored = self.project.project_directory / relative
        stored.parent.mkdir(parents=True, exist_ok=True)
        temporary = stored.with_name(f".{stored.name}.{uuid7()}.tmp")
        try:
            shutil.copyfile(source, temporary)
            with temporary.open("r+b") as handle:
                os.fsync(handle.fileno())
            os.replace(temporary, stored)
        finally:
            temporary.unlink(missing_ok=True)
        sha256, stored_size = self._file_digest(stored)
        try:
            preview = self._preview(stored, entity_type, mapping, sheet_name)
        except Exception:
            stored.unlink(missing_ok=True)
            raise
        summary = {
            "entityType": entity_type,
            "relativePath": relative,
            "sha256": sha256,
            "sizeBytes": stored_size,
            "mapping": preview["mapping"],
            "sheetName": preview["sheetName"],
            "rowCount": preview["rowCount"],
            "errorCount": len(preview["errors"]),
            "warningCount": len(preview["warnings"]),
        }
        now = utc_now()
        self.project.connection.execute(
            """
            INSERT INTO import_jobs(id, source_type, source_name, status, summary, created_at, updated_at)
            VALUES (?, ?, ?, 'preview', ?, ?, ?)
            """,
            (job_id, source.suffix.lower().lstrip("."), source.name[:255], json.dumps(summary, ensure_ascii=False, separators=(",", ":")), now, now),
        )
        self.project.connection.execute(
            "INSERT INTO attachments(id, kind, relative_path, sha256, size_bytes, created_at) VALUES (?, 'import_source', ?, ?, ?, ?)",
            (uuid7(), relative, sha256, stored_size, now),
        )
        return {"id": job_id, **{key: value for key, value in preview.items() if key != "allRecords"}}

    def remap_preview(
        self,
        job_id: str,
        *,
        mapping: dict[str, str],
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        job, summary, path = self._preview_job(job_id)
        preview = self._preview(path, summary["entityType"], mapping, sheet_name or summary.get("sheetName"))
        summary.update(
            {
                "mapping": preview["mapping"],
                "sheetName": preview["sheetName"],
                "rowCount": preview["rowCount"],
                "errorCount": len(preview["errors"]),
                "warningCount": len(preview["warnings"]),
            }
        )
        self.project.connection.execute(
            "UPDATE import_jobs SET summary = ?, updated_at = ? WHERE id = ?",
            (json.dumps(summary, ensure_ascii=False, separators=(",", ":")), utc_now(), job["id"]),
        )
        return {"id": job_id, **{key: value for key, value in preview.items() if key != "allRecords"}}

    def confirm_import(self, job_id: str, expected_revision: int) -> dict[str, Any]:
        job, summary, path = self._preview_job(job_id)
        if self.project.revision != expected_revision:
            from stt_desktop.storage.project import RevisionConflictError

            raise RevisionConflictError(expected_revision, self.project.revision)
        sha256, size = self._file_digest(path)
        if sha256 != summary["sha256"] or size != int(summary["sizeBytes"]):
            raise ProjectError("导入原文件与预览时的哈希或大小不一致")
        preview = self._preview(
            path,
            summary["entityType"],
            dict(summary["mapping"]),
            summary.get("sheetName"),
            preview_limit=MAX_IMPORT_ROWS,
        )
        if preview["errors"]:
            raise ProjectError("导入仍存在校验错误，请返回预览修正字段映射或数据")
        backup = BackupService(self.project, self.workspace).create_backup(reason="pre-import")
        records = preview["allRecords"]
        batches: dict[str, list[dict[str, Any]]] = {
            summary["entityType"]: records
        }
        if summary["entityType"] == "teaching_task":
            tasks: list[dict[str, Any]] = []
            lessons: list[dict[str, Any]] = []
            for row_index, record in enumerate(records, start=1):
                task_id = uuid7()
                task = {"id": task_id, **record}
                tasks.append(task)
                remaining = int(record["weekly_slots"])
                duration = int(record.get("duration_slots") or 1)
                lesson_index = 0
                while remaining > 0:
                    lesson_duration = min(duration, remaining)
                    lessons.append(
                        {
                            "id": uuid7(),
                            "teaching_task_id": task_id,
                            "lesson_index": lesson_index,
                            "duration_slots": lesson_duration,
                            "source_id": f"import:{job_id}:{row_index}:{lesson_index}",
                            "week_bits": record.get("week_bits") or "1" * 20,
                            "day_bits": record.get("day_bits") or "11111",
                            "label": f"第{lesson_index + 1}次课",
                            "enabled": 1,
                        }
                    )
                    remaining -= lesson_duration
                    lesson_index += 1
            batches = {"teaching_task": tasks, "task_lesson": lessons}
        counts, revision = self.project.bulk_insert_entities(
            batches, expected_revision=expected_revision
        )
        completed_summary = {
            **summary,
            "confirmedCount": counts[summary["entityType"]],
            "confirmedRevision": revision,
            "preImportBackupId": backup["id"],
            "generatedLessonCount": counts.get("task_lesson", 0),
        }
        self.project.connection.execute(
            "UPDATE import_jobs SET status = 'confirmed', summary = ?, updated_at = ? WHERE id = ?",
            (json.dumps(completed_summary, ensure_ascii=False, separators=(",", ":")), utc_now(), job_id),
        )
        return {
            "id": job_id,
            "status": "confirmed",
            "entityType": summary["entityType"],
            "importedCount": counts[summary["entityType"]],
            "generatedLessonCount": counts.get("task_lesson", 0),
            "revision": revision,
            "backupId": backup["id"],
        }

    def abandon(self, job_id: str) -> dict[str, Any]:
        cursor = self.project.connection.execute(
            "UPDATE import_jobs SET status = 'abandoned', updated_at = ? WHERE id = ? AND status = 'preview'",
            (utc_now(), job_id),
        )
        if cursor.rowcount != 1:
            raise ProjectError("待放弃的导入预览不存在")
        return {"id": job_id, "status": "abandoned"}

    def list_imports(self) -> list[dict[str, Any]]:
        rows = self.project.connection.execute(
            "SELECT * FROM import_jobs ORDER BY created_at DESC"
        ).fetchall()
        return [{**dict(row), "summary": json.loads(row["summary"] or "{}")} for row in rows]

    def create_template(
        self,
        *,
        entity_type: str,
        file_format: str,
        destination_path: str,
        overwrite: bool = False,
    ) -> dict[str, Any]:
        if entity_type not in IMPORT_SCHEMAS or file_format not in {"csv", "xlsx"}:
            raise ProjectError("导入模板类型或格式不受支持")
        suffix = f".{file_format}"
        destination = Path(destination_path).expanduser().resolve()
        if destination.suffix.lower() != suffix or not destination.parent.is_dir():
            raise ProjectError(f"模板目标必须是已存在目录中的 {suffix} 文件")
        if destination.exists() and not overwrite:
            raise ProjectError("模板目标已存在，必须明确确认覆盖")
        template_id = uuid7()
        relative = f"artifacts/exports/import-template-{entity_type}-{template_id}{suffix}"
        internal = self.project.project_directory / relative
        internal.parent.mkdir(parents=True, exist_ok=True)
        temporary = internal.with_name(f".{internal.name}.{uuid7()}.tmp")
        fields = IMPORT_SCHEMAS[entity_type]
        headers = [field.aliases[0] for field in fields]
        try:
            if file_format == "csv":
                with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
                    csv.writer(handle).writerow(headers)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "导入模板"
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.fill = PatternFill("solid", fgColor="1F604D")
                    cell.font = Font(color="FFFFFF", bold=True)
                guide = workbook.create_sheet("字段说明")
                guide.append(("字段", "内部字段", "必填"))
                for field in fields:
                    guide.append((field.aliases[0], field.name, "是" if field.required else "否"))
                workbook.save(temporary)
            with temporary.open("r+b") as handle:
                os.fsync(handle.fileno())
            os.replace(temporary, internal)
        finally:
            temporary.unlink(missing_ok=True)
        sha256, size = self._file_digest(internal)
        copied = self._copy_template(internal, destination, overwrite)
        now = utc_now()
        summary = {
            "entityType": entity_type,
            "template": True,
            "sha256": sha256,
            "sizeBytes": size,
        }
        self.project.connection.execute(
            """
            INSERT INTO export_jobs(id, export_type, status, relative_path, summary, created_at, updated_at)
            VALUES (?, ?, 'succeeded', ?, ?, ?, ?)
            """,
            (template_id, f"import_template_{file_format}", relative, json.dumps(summary, ensure_ascii=False, separators=(",", ":")), now, now),
        )
        self.project.connection.execute(
            "INSERT INTO artifacts(id, kind, relative_path, sha256, size_bytes, created_at) VALUES (?, 'import_template', ?, ?, ?, ?)",
            (uuid7(), relative, sha256, size, now),
        )
        return {
            "id": template_id,
            "entityType": entity_type,
            "format": file_format,
            "destinationPath": copied,
            "sha256": sha256,
            "sizeBytes": size,
        }

    def _preview(
        self,
        path: Path,
        entity_type: str,
        mapping: dict[str, str] | None,
        sheet_name: str | None,
        *,
        preview_limit: int = 50,
    ) -> dict[str, Any]:
        headers, raw_rows, resolved_sheet, sheets = self._read_table(path, sheet_name)
        resolved_mapping = self._resolve_mapping(entity_type, headers, mapping)
        records, errors, warnings = self._normalize_rows(entity_type, headers, raw_rows, resolved_mapping)
        return {
            "entityType": entity_type,
            "headers": headers,
            "mapping": resolved_mapping,
            "sheetName": resolved_sheet,
            "availableSheets": sheets,
            "rowCount": len(raw_rows),
            "previewRows": records[:preview_limit],
            "allRecords": records,
            "errors": errors[:200],
            "warnings": warnings[:200],
            "truncatedIssues": len(errors) > 200 or len(warnings) > 200,
            "canConfirm": not errors and bool(records),
        }

    def _read_table(
        self, path: Path, sheet_name: str | None
    ) -> tuple[list[str], list[list[Any]], str | None, list[str]]:
        if path.suffix.lower() == ".csv":
            text = self._read_csv_text(path)
            sample = text[:8192]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.reader(text.splitlines(), dialect))
            return self._clean_table(rows, None, [])
        self._inspect_xlsx_archive(path)
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheets = workbook.sheetnames
            if not sheets:
                raise ProjectError("XLSX 不包含工作表")
            selected = sheet_name or sheets[0]
            if selected not in sheets:
                raise ProjectError("指定的 XLSX 工作表不存在")
            sheet = workbook[selected]
            rows: list[list[Any]] = []
            for row_index, cells in enumerate(sheet.iter_rows(), start=1):
                if row_index > MAX_IMPORT_ROWS + 1:
                    raise ProjectError("导入数据超过 100,000 行限制")
                values = []
                for cell in cells[:MAX_IMPORT_COLUMNS]:
                    if cell.data_type == "f":
                        raise ProjectError(f"XLSX 第 {row_index} 行包含公式，导入文件只允许静态值")
                    values.append(cell.value)
                rows.append(values)
            return self._clean_table(rows, selected, sheets)
        finally:
            workbook.close()

    @staticmethod
    def _clean_table(
        rows: list[list[Any]], sheet_name: str | None, sheets: list[str]
    ) -> tuple[list[str], list[list[Any]], str | None, list[str]]:
        while rows and not any(_text(value) for value in rows[0]):
            rows.pop(0)
        if not rows:
            raise ProjectError("导入文件没有表头")
        headers = [_text(value) for value in rows[0][:MAX_IMPORT_COLUMNS]]
        if not headers or any(not value for value in headers):
            raise ProjectError("导入表头不能为空")
        if len(set(headers)) != len(headers):
            raise ProjectError("导入表头存在重复名称")
        data = [row[:len(headers)] for row in rows[1:] if any(_text(value) for value in row)]
        if len(data) > MAX_IMPORT_ROWS:
            raise ProjectError("导入数据超过 100,000 行限制")
        return headers, data, sheet_name, sheets

    @staticmethod
    def _read_csv_text(path: Path) -> str:
        content = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ProjectError("CSV 编码无法识别，请保存为 UTF-8 或 GB18030")

    @staticmethod
    def _resolve_mapping(
        entity_type: str, headers: list[str], mapping: dict[str, str] | None
    ) -> dict[str, str]:
        fields = IMPORT_SCHEMAS[entity_type]
        allowed = {field.name for field in fields}
        if mapping is not None:
            unknown_headers = set(mapping) - set(headers)
            unknown_targets = {target for target in mapping.values() if target and target not in allowed}
            if unknown_headers or unknown_targets:
                raise ProjectError("字段映射包含未知表头或目标字段")
            resolved = {header: target for header, target in mapping.items() if target}
        else:
            aliases = {
                alias.strip().lower(): field.name
                for field in fields
                for alias in (field.name, *field.aliases)
            }
            resolved = {
                header: aliases[header.strip().lower()]
                for header in headers
                if header.strip().lower() in aliases
            }
        targets = list(resolved.values())
        if len(targets) != len(set(targets)):
            raise ProjectError("多个源字段不能映射到同一目标字段")
        missing = [field.name for field in fields if field.required and field.name not in targets]
        if missing:
            raise ProjectError(f"字段映射缺少必要字段: {', '.join(missing)}")
        return resolved

    def _normalize_rows(
        self,
        entity_type: str,
        headers: list[str],
        rows: list[list[Any]],
        mapping: dict[str, str],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        fields = {field.name: field for field in IMPORT_SCHEMAS[entity_type]}
        records: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        seen_records: set[tuple[Any, ...]] = set()
        for row_number, values in enumerate(rows, start=2):
            source = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
            record: dict[str, Any] = {}
            row_errors: list[str] = []
            for header, target in mapping.items():
                value = source.get(header)
                if len(_text(value)) > MAX_CELL_LENGTH:
                    row_errors.append(f"{header} 超过 {MAX_CELL_LENGTH} 字符")
                    continue
                field = fields[target]
                try:
                    converted = field.converter(value) if field.converter else value
                except (TypeError, ValueError) as exc:
                    row_errors.append(f"{header}: {exc}")
                    continue
                if converted is not None:
                    record[target] = converted
            for field in fields.values():
                if field.required and not _text(record.get(field.name)):
                    row_errors.append(f"{field.name} 不能为空")
            if row_errors:
                errors.append({"row": row_number, "messages": row_errors})
                continue
            try:
                prepared = self._resolve_references(entity_type, record)
                duplicate_key = self._duplicate_key(entity_type, prepared)
                if duplicate_key in seen_records:
                    raise ProjectError("记录在导入文件内重复")
                seen_records.add(duplicate_key)
                self._check_existing_duplicate(entity_type, prepared)
            except ProjectError as exc:
                errors.append({"row": row_number, "messages": [str(exc)]})
                continue
            records.append(prepared)
            unmapped_values = [header for header in headers if header not in mapping and _text(source.get(header))]
            if unmapped_values:
                warnings.append({"row": row_number, "message": f"未导入字段: {', '.join(unmapped_values)}"})
        return records, errors, warnings

    def _resolve_references(self, entity_type: str, record: dict[str, Any]) -> dict[str, Any]:
        result = dict(record)
        references = {
            "room": {"room_type_name": ("room_types", "room_type_id")},
            "homeroom": {
                "grade_name": ("grades", "grade_id"),
                "term_name": ("terms", "term_id"),
                "head_teacher_name": ("teachers", "head_teacher_id"),
                "default_room_name": ("rooms", "default_room_id"),
            },
            "course_plan": {
                "term_name": ("terms", "term_id"),
                "homeroom_name": ("homerooms", "homeroom_id"),
                "subject_name": ("subjects", "subject_id"),
            },
            "teaching_task": {
                "term_name": ("terms", "term_id"),
                "homeroom_name": ("homerooms", "homeroom_id"),
                "subject_name": ("subjects", "subject_id"),
                "primary_teacher_name": ("teachers", "primary_teacher_id"),
                "required_room_type_name": ("room_types", "required_room_type"),
                "fixed_room_name": ("rooms", "fixed_room_id"),
            },
        }
        for source_field, (table, target_field) in references.get(entity_type, {}).items():
            name = result.pop(source_field, None)
            if not name:
                continue
            rows = self.project.connection.execute(
                f"SELECT id FROM {table} WHERE name = ? ORDER BY id",  # noqa: S608 - fixed allowlist
                (name,),
            ).fetchall()
            if len(rows) != 1:
                raise ProjectError(f"引用 {name} 在本地 {table} 中不存在或不唯一")
            result[target_field] = rows[0]["id"]
        if entity_type in {"course_plan", "teaching_task"} and not result.get(
            "term_id"
        ):
            homeroom = self.project.connection.execute(
                "SELECT term_id FROM homerooms WHERE id = ?",
                (result["homeroom_id"],),
            ).fetchone()
            if homeroom and homeroom["term_id"]:
                result["term_id"] = homeroom["term_id"]
        if entity_type == "teaching_task":
            term_id = result.get("term_id")
            rows = self.project.connection.execute(
                """
                SELECT id FROM course_plans
                WHERE homeroom_id = ? AND subject_id = ?
                  AND ((term_id IS NULL AND ? IS NULL) OR term_id = ?)
                ORDER BY id
                """,
                (
                    result["homeroom_id"],
                    result["subject_id"],
                    term_id,
                    term_id,
                ),
            ).fetchall()
            if len(rows) == 1:
                result["course_plan_id"] = rows[0]["id"]
            elif len(rows) > 1:
                raise ProjectError("匹配到多个课程计划，无法自动关联")
        return result

    def _check_existing_duplicate(self, entity_type: str, record: dict[str, Any]) -> None:
        table = {
            "teacher": "teachers",
            "subject": "subjects",
            "grade": "grades",
            "room_type": "room_types",
            "room": "rooms",
            "homeroom": "homerooms",
        }.get(entity_type)
        if table:
            row = self.project.connection.execute(
                f"SELECT id FROM {table} WHERE name = ? LIMIT 1",  # noqa: S608 - fixed allowlist
                (record["name"],),
            ).fetchone()
            label = record["name"]
        elif entity_type == "course_plan":
            term_id = record.get("term_id")
            row = self.project.connection.execute(
                """
                SELECT id FROM course_plans
                WHERE homeroom_id = ? AND subject_id = ?
                  AND ((term_id IS NULL AND ? IS NULL) OR term_id = ?)
                LIMIT 1
                """,
                (record["homeroom_id"], record["subject_id"], term_id, term_id),
            ).fetchone()
            label = "同学期、班级和科目的课程计划"
        elif entity_type == "teaching_task":
            term_id = record.get("term_id")
            row = self.project.connection.execute(
                """
                SELECT id FROM teaching_tasks
                WHERE homeroom_id = ? AND subject_id = ?
                  AND primary_teacher_id = ?
                  AND ((term_id IS NULL AND ? IS NULL) OR term_id = ?)
                LIMIT 1
                """,
                (
                    record["homeroom_id"],
                    record["subject_id"],
                    record.get("primary_teacher_id"),
                    term_id,
                    term_id,
                ),
            ).fetchone()
            label = "同学期、班级、科目和教师的教学任务"
        else:
            raise ProjectError("导入类型不受支持")
        if row:
            raise ProjectError(f"本地已存在{label}")

    @staticmethod
    def _duplicate_key(entity_type: str, record: dict[str, Any]) -> tuple[Any, ...]:
        if entity_type == "course_plan":
            return (
                record.get("term_id"),
                record.get("homeroom_id"),
                record.get("subject_id"),
            )
        if entity_type == "teaching_task":
            return (
                record.get("term_id"),
                record.get("homeroom_id"),
                record.get("subject_id"),
                record.get("primary_teacher_id"),
            )
        return (record.get("name"),)

    def _preview_job(self, job_id: str) -> tuple[dict[str, Any], dict[str, Any], Path]:
        row = self.project.connection.execute(
            "SELECT * FROM import_jobs WHERE id = ? AND status = 'preview'", (job_id,)
        ).fetchone()
        if not row:
            raise ProjectError("导入预览不存在或已结束")
        job = dict(row)
        summary = json.loads(job["summary"])
        path = (self.project.project_directory / summary["relativePath"]).resolve()
        if self.project.project_directory not in path.parents or not path.is_file():
            raise ProjectError("导入附件路径无效或文件缺失")
        return job, summary, path

    @staticmethod
    def _inspect_xlsx_archive(path: Path) -> None:
        try:
            with zipfile.ZipFile(path, "r") as archive:
                infos = archive.infolist()
                if len(infos) > 2_000:
                    raise ProjectError("XLSX 内部文件数量超过限制")
                total = 0
                for info in infos:
                    name = PurePosixPath(info.filename)
                    if name.is_absolute() or ".." in name.parts or "\\" in info.filename:
                        raise ProjectError("XLSX 包含非法内部路径")
                    total += info.file_size
                    if total > 200 * 1024 * 1024:
                        raise ProjectError("XLSX 解压体积超过 200 MiB")
                    if info.compress_size and info.file_size / info.compress_size > 1_000:
                        raise ProjectError("XLSX 包含异常压缩比条目")
        except zipfile.BadZipFile as exc:
            raise ProjectError("XLSX 文件结构损坏") from exc

    @staticmethod
    def _file_digest(path: Path) -> tuple[str, int]:
        digest = hashlib.sha256()
        size = 0
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
                size += len(chunk)
        return digest.hexdigest(), size

    @staticmethod
    def _copy_template(source: Path, destination: Path, overwrite: bool) -> str:
        if destination.exists() and not overwrite:
            raise ProjectError("模板目标已存在，必须明确确认覆盖")
        temporary = destination.with_name(f".{destination.name}.{uuid7()}.tmp")
        try:
            shutil.copyfile(source, temporary)
            with temporary.open("r+b") as handle:
                os.fsync(handle.fileno())
            os.replace(temporary, destination)
        finally:
            temporary.unlink(missing_ok=True)
        return str(destination)
