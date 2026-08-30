from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from stt_desktop.scheduling import SchedulingService
from stt_desktop.storage.project import ProjectError, ProjectRepository, utc_now, uuid7


EXPORT_SUFFIXES = {
    "csv": ".csv",
    "xlsx": ".xlsx",
    "pdf": ".pdf",
    "problem_xml": ".xml",
    "solution_xml": ".xml",
}
HEADERS = ("星期", "开始课节", "连续课时", "班级", "课程", "教师", "教室", "周次位图", "课次ID")


class ExportService:
    def __init__(self, project: ProjectRepository) -> None:
        self.project = project

    def export_candidate(
        self,
        *,
        candidate_id: str,
        export_type: str,
        destination_path: str | None = None,
        overwrite: bool = False,
    ) -> dict[str, Any]:
        if export_type not in EXPORT_SUFFIXES:
            raise ProjectError("导出类型只支持 CSV、XLSX、PDF、Problem XML 或 Solution XML")
        candidate = self._candidate(candidate_id)
        if int(candidate["hard_violations"]) != 0:
            raise ProjectError("只有硬约束零违例候选可以导出")
        job_id = uuid7()
        suffix = EXPORT_SUFFIXES[export_type]
        relative = f"artifacts/exports/{job_id}{suffix}"
        target = self.project.project_directory / relative
        now = utc_now()
        self.project.connection.execute(
            """
            INSERT INTO export_jobs(id, export_type, status, relative_path, summary, created_at, updated_at)
            VALUES (?, ?, 'running', ?, '{}', ?, ?)
            """,
            (job_id, export_type, relative, now, now),
        )
        try:
            rows = self._timetable_rows(candidate_id)
            writers: dict[str, Callable[[Path, dict, list[dict]], None]] = {
                "csv": self._write_csv,
                "xlsx": self._write_xlsx,
                "pdf": self._write_pdf,
                "problem_xml": self._write_problem_xml,
                "solution_xml": self._write_solution_xml,
            }
            target.parent.mkdir(parents=True, exist_ok=True)
            temporary = target.with_name(f".{target.name}.{uuid7()}.tmp")
            try:
                writers[export_type](temporary, candidate, rows)
                with temporary.open("r+b") as handle:
                    os.fsync(handle.fileno())
                temporary.replace(target)
            finally:
                temporary.unlink(missing_ok=True)
            sha256, size = self._file_digest(target)
            destination = None
            if destination_path:
                destination = self._copy_to_destination(target, destination_path, suffix, overwrite)
            summary = {
                "candidateId": candidate_id,
                "rowCount": len(rows),
                "sha256": sha256,
                "sizeBytes": size,
                "copiedToUserDestination": destination is not None,
            }
            finished = utc_now()
            self.project.connection.execute(
                "UPDATE export_jobs SET status = 'succeeded', summary = ?, updated_at = ? WHERE id = ?",
                (json.dumps(summary, ensure_ascii=False, separators=(",", ":")), finished, job_id),
            )
            self.project.connection.execute(
                "INSERT INTO artifacts(id, kind, relative_path, sha256, size_bytes, created_at) VALUES (?, 'export', ?, ?, ?, ?)",
                (uuid7(), relative, sha256, size, finished),
            )
            return {
                "id": job_id,
                "exportType": export_type,
                "status": "succeeded",
                "relativePath": relative,
                "fileName": target.name,
                "destinationPath": destination,
                **summary,
            }
        except Exception as exc:
            failed = utc_now()
            summary = {"code": "EXPORT_FAILED", "message": str(exc)[:500]}
            self.project.connection.execute(
                "UPDATE export_jobs SET status = 'failed', summary = ?, updated_at = ? WHERE id = ?",
                (json.dumps(summary, ensure_ascii=False, separators=(",", ":")), failed, job_id),
            )
            raise

    def list_exports(self) -> list[dict[str, Any]]:
        rows = self.project.connection.execute(
            "SELECT * FROM export_jobs ORDER BY created_at DESC"
        ).fetchall()
        return [
            {**dict(row), "summary": json.loads(row["summary"] or "{}")}
            for row in rows
        ]

    def _candidate(self, candidate_id: str) -> dict[str, Any]:
        row = self.project.connection.execute(
            "SELECT * FROM candidates WHERE id = ? AND status = 'valid'", (candidate_id,)
        ).fetchone()
        if not row:
            raise ProjectError("要导出的候选不存在或已失效")
        return dict(row)

    def _timetable_rows(self, candidate_id: str) -> list[dict[str, Any]]:
        rows = self.project.connection.execute(
            """
            SELECT e.weekday, e.start_slot, e.duration_slots, e.week_bits,
                   e.task_lesson_id, h.name AS homeroom_name, s.name AS subject_name,
                   t.name AS teacher_name, r.name AS room_name
            FROM timetable_entries e
            LEFT JOIN homerooms h ON h.id = e.homeroom_id
            LEFT JOIN subjects s ON s.id = e.subject_id
            LEFT JOIN teachers t ON t.id = e.teacher_id
            LEFT JOIN rooms r ON r.id = e.room_id
            WHERE e.candidate_id = ?
            ORDER BY e.weekday, e.start_slot, h.name, s.name
            """,
            (candidate_id,),
        ).fetchall()
        return [dict(row) for row in rows]

    @staticmethod
    def _row_values(row: dict[str, Any]) -> tuple[Any, ...]:
        return (
            f"星期{row['weekday']}",
            row["start_slot"],
            row["duration_slots"],
            row["homeroom_name"] or "",
            row["subject_name"] or "",
            row["teacher_name"] or "",
            row["room_name"] or "",
            row["week_bits"],
            row["task_lesson_id"] or "",
        )

    @classmethod
    def _spreadsheet_values(cls, row: dict[str, Any]) -> tuple[Any, ...]:
        return tuple(
            f"'{value}" if isinstance(value, str) and value.startswith(("=", "+", "-", "@")) else value
            for value in cls._row_values(row)
        )

    def _write_csv(self, path: Path, _: dict, rows: list[dict]) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(HEADERS)
            writer.writerows(self._spreadsheet_values(row) for row in rows)

    def _write_xlsx(self, path: Path, candidate: dict, rows: list[dict]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "候选课表"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:I{max(1, len(rows) + 1)}"
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(self._spreadsheet_values(row))
        header_fill = PatternFill("solid", fgColor="1F604D")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = (12, 12, 12, 18, 18, 18, 18, 28, 38)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        summary = workbook.create_sheet("候选信息")
        summary.append(("项目", self.project.project_info()["name"]))
        summary.append(("候选 ID", candidate["id"]))
        summary.append(("总得分", candidate["total_score"]))
        summary.append(("硬约束违例", candidate["hard_violations"]))
        summary.append(("生成时间", candidate["created_at"]))
        summary.column_dimensions["A"].width = 22
        summary.column_dimensions["B"].width = 56
        workbook.save(path)

    def _write_pdf(self, path: Path, candidate: dict, rows: list[dict]) -> None:
        font_name = self._register_cjk_font()
        document = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
            title=f"{self.project.project_info()['name']} 候选课表",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "KariosTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=16,
            leading=21,
            alignment=TA_CENTER,
        )
        body_style = ParagraphStyle(
            "KariosBody", parent=styles["BodyText"], fontName=font_name, fontSize=7, leading=9
        )
        story: list[Any] = [
            Paragraph(f"{self.project.project_info()['name']} · 候选课表", title_style),
            Paragraph(
                f"候选 {candidate['id']}　总得分 {candidate['total_score']}　导出时间 {utc_now()}",
                body_style,
            ),
            Spacer(1, 4 * mm),
        ]
        table_data = [[Paragraph(str(value), body_style) for value in HEADERS[:-2]]]
        for row in rows:
            table_data.append(
                [Paragraph(str(value), body_style) for value in self._row_values(row)[:-2]]
            )
        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[22 * mm, 18 * mm, 18 * mm, 34 * mm, 34 * mm, 34 * mm, 34 * mm],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F604D")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9C9C2")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7F4")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)

        def footer(canvas, doc) -> None:
            canvas.saveState()
            canvas.setFont(font_name, 7)
            canvas.setFillColor(colors.HexColor("#61756D"))
            canvas.drawString(10 * mm, 6 * mm, "时奕教务排课 · 本地导出")
            canvas.drawRightString(landscape(A4)[0] - 10 * mm, 6 * mm, f"第 {doc.page} 页")
            canvas.restoreState()

        document.build(story, onFirstPage=footer, onLaterPages=footer)

    def _write_problem_xml(self, path: Path, candidate: dict, _: list[dict]) -> None:
        source = self.project.project_directory / f"artifacts/problem/{candidate['round_id']}.xml"
        if not source.is_file():
            parent_round = self.project.connection.execute(
                "SELECT parent_candidate_id FROM scheduling_rounds WHERE id = ?", (candidate["round_id"],)
            ).fetchone()
            if parent_round and parent_round["parent_candidate_id"]:
                parent = self._candidate(parent_round["parent_candidate_id"])
                source = self.project.project_directory / f"artifacts/problem/{parent['round_id']}.xml"
        if not source.is_file():
            snapshot = self._snapshot(candidate["snapshot_id"])
            xml, _ = SchedulingService(self.project)._build_problem(snapshot)
            path.write_text(xml, encoding="utf-8", newline="\n")
            return
        shutil.copyfile(source, path)

    def _write_solution_xml(self, path: Path, candidate: dict, _: list[dict]) -> None:
        diagnostics = json.loads(candidate["diagnostics"] or "{}")
        relative = str(diagnostics.get("solutionPath") or "")
        source = (self.project.project_directory / relative).resolve()
        if not relative or self.project.project_directory not in source.parents or not source.is_file():
            raise ProjectError("候选 Solution XML 不存在或路径无效")
        shutil.copyfile(source, path)

    def _snapshot(self, snapshot_id: str) -> dict[str, Any]:
        row = self.project.connection.execute(
            "SELECT payload_path FROM data_snapshots WHERE id = ?", (snapshot_id,)
        ).fetchone()
        if not row:
            raise ProjectError("候选数据快照不存在")
        source = (self.project.project_directory / row["payload_path"]).resolve()
        if self.project.project_directory not in source.parents:
            raise ProjectError("候选数据快照路径越界")
        return json.loads(source.read_text(encoding="utf-8"))

    @staticmethod
    def _register_cjk_font() -> str:
        existing = set(pdfmetrics.getRegisteredFontNames())
        if "KariosCJK" in existing:
            return "KariosCJK"
        candidates = (
            Path("C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
            Path("/System/Library/Fonts/PingFang.ttc"),
        )
        for candidate in candidates:
            if not candidate.is_file():
                continue
            try:
                pdfmetrics.registerFont(TTFont("KariosCJK", str(candidate), subfontIndex=0))
                return "KariosCJK"
            except Exception:
                continue
        raise ProjectError("未找到可用于中文 PDF 的系统字体（微软雅黑、黑体、Noto Sans CJK 或苹方）")

    @staticmethod
    def _file_digest(path: Path) -> tuple[str, int]:
        digest = hashlib.sha256()
        size = 0
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
                size += len(chunk)
        return digest.hexdigest(), size

    @staticmethod
    def _copy_to_destination(source: Path, destination_path: str, suffix: str, overwrite: bool) -> str:
        destination = Path(destination_path).expanduser().resolve()
        if destination.suffix.lower() != suffix:
            raise ProjectError(f"目标文件扩展名必须是 {suffix}")
        if not destination.parent.is_dir():
            raise ProjectError("目标目录不存在")
        if destination.exists() and not overwrite:
            raise ProjectError("目标文件已存在，必须明确确认覆盖")
        temporary = destination.with_name(f".{destination.name}.{uuid7()}.tmp")
        try:
            shutil.copyfile(source, temporary)
            with temporary.open("r+b") as handle:
                os.fsync(handle.fileno())
            os.replace(temporary, destination)
        finally:
            temporary.unlink(missing_ok=True)
        return str(destination)
