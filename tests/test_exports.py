from __future__ import annotations

from pathlib import Path
from xml.etree.ElementTree import fromstring

import pytest
from openpyxl import load_workbook

from stt_desktop.scheduling import SchedulingService
from stt_desktop.storage import ProjectError, ProjectWorkspace
from stt_desktop.transfers import ExportService


def export_fixture(tmp_path: Path):
    project = ProjectWorkspace(tmp_path / "workspace").create_project("中文课表项目")
    _, revision = project.bulk_insert_entities(
        {
            "term": [{"id": "term-1", "name": "第一学期"}],
            "bell_schedule": [
                {
                    "id": "schedule-1",
                    "term_id": "term-1",
                    "name": "默认作息",
                    "is_default": 1,
                }
            ],
            "time_slot": [
                {
                    "id": f"slot-{index}",
                    "bell_schedule_id": "schedule-1",
                    "weekday": 1,
                    "period_index": index,
                    "label": f"第{index + 1}节",
                    "start_slot": index,
                    "length_slots": 1,
                    "start_time_minutes": 480 + index * 50,
                    "end_time_minutes": 520 + index * 50,
                }
                for index in range(2)
            ],
            "teacher": [{"id": "teacher-1", "name": "张老师"}],
            "subject": [{"id": "subject-1", "name": "数学"}],
            "homeroom": [
                {"id": "homeroom-1", "term_id": "term-1", "name": "一年级一班"}
            ],
        },
        0,
    )
    project.save_teaching_task_bundle(
        {
            "term_id": "term-1",
            "homeroom_id": "homeroom-1",
            "subject_id": "subject-1",
            "primary_teacher_id": "teacher-1",
            "weekly_slots": 2,
            "duration_slots": 1,
        },
        revision,
    )
    candidate = SchedulingService(project).run_round(time_budget_seconds=10)
    return project, candidate["candidate_id"]


def test_candidate_exports_csv_xlsx_pdf_and_xml(tmp_path: Path) -> None:
    project, candidate_id = export_fixture(tmp_path)
    try:
        service = ExportService(project)
        exported = {
            kind: service.export_candidate(candidate_id=candidate_id, export_type=kind)
            for kind in ("csv", "xlsx", "pdf", "problem_xml", "solution_xml")
        }
        paths = {
            kind: project.project_directory / result["relativePath"]
            for kind, result in exported.items()
        }
        assert paths["csv"].read_bytes().startswith(b"\xef\xbb\xbf")
        assert "一年级一班" in paths["csv"].read_text(encoding="utf-8-sig")

        workbook = load_workbook(paths["xlsx"], read_only=True)
        assert workbook.sheetnames == ["候选课表", "候选信息"]
        assert workbook["候选课表"]["D2"].value == "一年级一班"
        workbook.close()

        assert paths["pdf"].read_bytes().startswith(b"%PDF")
        assert fromstring(paths["problem_xml"].read_text(encoding="utf-8")).tag == "problem"
        assert fromstring(paths["solution_xml"].read_text(encoding="utf-8")).tag == "solution"
        assert len(service.list_exports()) == 5
        assert all(item["status"] == "succeeded" for item in service.list_exports())
    finally:
        project.close()


def test_export_to_user_destination_is_atomic_and_requires_explicit_overwrite(
    tmp_path: Path,
) -> None:
    project, candidate_id = export_fixture(tmp_path)
    try:
        service = ExportService(project)
        destination = tmp_path / "导出课表.csv"
        result = service.export_candidate(
            candidate_id=candidate_id,
            export_type="csv",
            destination_path=str(destination),
        )
        assert result["destinationPath"] == str(destination.resolve())
        original = destination.read_bytes()

        with pytest.raises(ProjectError, match="明确确认覆盖"):
            service.export_candidate(
                candidate_id=candidate_id,
                export_type="csv",
                destination_path=str(destination),
            )
        assert destination.read_bytes() == original

        service.export_candidate(
            candidate_id=candidate_id,
            export_type="csv",
            destination_path=str(destination),
            overwrite=True,
        )
        assert destination.read_bytes().startswith(b"\xef\xbb\xbf")

        with pytest.raises(ProjectError, match="扩展名"):
            service.export_candidate(
                candidate_id=candidate_id,
                export_type="xlsx",
                destination_path=str(tmp_path / "wrong.csv"),
            )
    finally:
        project.close()
