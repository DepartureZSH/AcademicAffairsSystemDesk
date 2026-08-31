from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from stt_desktop.storage import ProjectError, ProjectWorkspace, RevisionConflictError
from stt_desktop.transfers import ImportService


def write_csv(path: Path, rows: list[list[str]], encoding: str = "utf-8-sig") -> None:
    with path.open("w", encoding=encoding, newline="") as handle:
        csv.writer(handle).writerows(rows)


def test_csv_preview_mapping_confirm_and_pre_import_backup(tmp_path: Path) -> None:
    workspace = ProjectWorkspace(tmp_path / "workspace")
    project = workspace.create_project("CSV 导入")
    source = tmp_path / "教师名单.csv"
    write_csv(
        source,
        [
            ["教师工号", "教师姓名", "部门", "状态", "忽略备注"],
            ["T001", "张老师", "数学组", "在职", "本地备注"],
            ["T002", "李老师", "语文组", "启用", ""],
        ],
    )
    try:
        service = ImportService(project, workspace)
        preview = service.preview_file(source_path=str(source), entity_type="teacher")
        assert preview["mapping"] == {
            "教师工号": "employee_no",
            "教师姓名": "name",
            "部门": "department",
            "状态": "status",
        }
        assert preview["canConfirm"]
        assert preview["rowCount"] == 2
        assert "allRecords" not in preview
        assert preview["warnings"][0]["message"] == "未导入字段: 忽略备注"
        assert project.list_entities("teacher") == []

        result = service.confirm_import(preview["id"], expected_revision=0)
        assert result["status"] == "confirmed"
        assert result["importedCount"] == 2
        assert project.revision == 1
        assert {item["name"] for item in project.list_entities("teacher")} == {"李老师", "张老师"}
        backups = project.connection.execute(
            "SELECT * FROM backup_records WHERE id = ?", (result["backupId"],)
        ).fetchall()
        assert len(backups) == 1
        assert backups[0]["reason"] == "pre-import"
        assert service.list_imports()[0]["status"] == "confirmed"
    finally:
        project.close()


def test_xlsx_custom_mapping_and_reference_resolution(tmp_path: Path) -> None:
    workspace = ProjectWorkspace(tmp_path / "workspace")
    project = workspace.create_project("XLSX 导入")
    project.save_entity("room_type", {"name": "实验室"}, 0)
    source = tmp_path / "教室.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "教室数据"
    sheet.append(["房间", "编号", "人数", "类别"])
    sheet.append(["物理实验室", "A101", 48, "实验室"])
    workbook.save(source)
    try:
        service = ImportService(project, workspace)
        initial = service.preview_file(
            source_path=str(source),
            entity_type="room",
            mapping={"房间": "name", "编号": "room_no", "人数": "capacity"},
            sheet_name="教室数据",
        )
        assert initial["canConfirm"]
        remapped = service.remap_preview(
            initial["id"],
            mapping={
                "房间": "name",
                "编号": "room_no",
                "人数": "capacity",
                "类别": "room_type_name",
            },
        )
        assert remapped["previewRows"][0]["room_type_id"]
        result = service.confirm_import(initial["id"], expected_revision=1)
        assert result["importedCount"] == 1
        room = project.list_entities("room")[0]
        assert room["capacity"] == 48
        assert room["room_type_id"] == project.list_entities("room_type")[0]["id"]
    finally:
        project.close()


def test_import_errors_formula_tampering_and_revision_conflict(tmp_path: Path) -> None:
    workspace = ProjectWorkspace(tmp_path / "workspace")
    project = workspace.create_project("错误导入")
    duplicate = tmp_path / "duplicate.csv"
    write_csv(duplicate, [["姓名"], ["同名"], ["同名"]])
    try:
        service = ImportService(project, workspace)
        duplicate_preview = service.preview_file(
            source_path=str(duplicate), entity_type="teacher"
        )
        assert not duplicate_preview["canConfirm"]
        assert duplicate_preview["errors"][0]["row"] == 3
        with pytest.raises(ProjectError, match="校验错误"):
            service.confirm_import(duplicate_preview["id"], expected_revision=0)

        formula = tmp_path / "formula.xlsx"
        workbook = Workbook()
        workbook.active.append(["姓名"])
        workbook.active.append(["=HYPERLINK(\"https://invalid.example\",\"恶意\")"])
        workbook.save(formula)
        with pytest.raises(ProjectError, match="公式"):
            service.preview_file(source_path=str(formula), entity_type="teacher")

        valid = tmp_path / "valid.csv"
        write_csv(valid, [["姓名"], ["王老师"]])
        preview = service.preview_file(source_path=str(valid), entity_type="teacher")
        project.save_entity("subject", {"name": "数学"}, 0)
        with pytest.raises(RevisionConflictError):
            service.confirm_import(preview["id"], expected_revision=0)

        job = next(item for item in service.list_imports() if item["id"] == preview["id"])
        stored = project.project_directory / job["summary"]["relativePath"]
        stored.write_text("姓名\n篡改", encoding="utf-8")
        with pytest.raises(ProjectError, match="哈希或大小"):
            service.confirm_import(preview["id"], expected_revision=1)

        fractional = tmp_path / "fractional.csv"
        write_csv(fractional, [["教室名称", "容量"], ["阶梯教室", "48.5"]])
        fractional_preview = service.preview_file(
            source_path=str(fractional), entity_type="room"
        )
        assert not fractional_preview["canConfirm"]
        assert "容量: 不是有限整数" in fractional_preview["errors"][0]["messages"]
    finally:
        project.close()


def test_localized_csv_and_xlsx_import_templates(tmp_path: Path) -> None:
    workspace = ProjectWorkspace(tmp_path / "workspace")
    project = workspace.create_project("模板")
    try:
        service = ImportService(project, workspace)
        csv_path = tmp_path / "教师模板.csv"
        xlsx_path = tmp_path / "班级模板.xlsx"
        csv_result = service.create_template(
            entity_type="teacher",
            file_format="csv",
            destination_path=str(csv_path),
        )
        xlsx_result = service.create_template(
            entity_type="homeroom",
            file_format="xlsx",
            destination_path=str(xlsx_path),
        )
        assert csv_result["destinationPath"] == str(csv_path.resolve())
        assert csv_path.read_text(encoding="utf-8-sig").startswith("工号,姓名,部门,状态")
        workbook = load_workbook(xlsx_path, read_only=True)
        assert workbook.sheetnames == ["导入模板", "字段说明"]
        assert workbook["导入模板"]["A1"].value == "班级名称"
        workbook.close()
        assert xlsx_result["sizeBytes"] > 0
    finally:
        project.close()


def test_course_plan_and_teaching_task_import_generate_lessons_atomically(
    tmp_path: Path,
) -> None:
    workspace = ProjectWorkspace(tmp_path / "workspace")
    project = workspace.create_project("计划任务导入")
    try:
        term, _ = project.save_entity(
            "term",
            {"name": "第一学期", "week_count": 20, "day_count": 5},
            project.revision,
        )
        project.save_entity(
            "teacher", {"name": "张老师"}, project.revision
        )
        project.save_entity("subject", {"name": "数学"}, project.revision)
        project.save_entity(
            "homeroom",
            {"name": "一年级一班", "term_id": term["id"]},
            project.revision,
        )
        service = ImportService(project, workspace)
        plan_file = tmp_path / "课程计划.csv"
        write_csv(
            plan_file,
            [
                ["班级", "科目", "每周课时", "连续课时", "允许连堂"],
                ["一年级一班", "数学", "5", "2", "是"],
            ],
        )
        plan_preview = service.preview_file(
            source_path=str(plan_file), entity_type="course_plan"
        )
        assert plan_preview["canConfirm"]
        plan_result = service.confirm_import(
            plan_preview["id"], expected_revision=project.revision
        )
        assert plan_result["importedCount"] == 1
        plan = project.list_entities("course_plan")[0]
        assert plan["term_id"] == term["id"]
        assert plan["duration_slots"] == 2

        task_file = tmp_path / "教学任务.csv"
        write_csv(
            task_file,
            [
                ["班级", "科目", "主讲教师", "每周课时", "连续课时"],
                ["一年级一班", "数学", "张老师", "5", "2"],
            ],
        )
        task_preview = service.preview_file(
            source_path=str(task_file), entity_type="teaching_task"
        )
        assert task_preview["canConfirm"]
        task_result = service.confirm_import(
            task_preview["id"], expected_revision=project.revision
        )

        assert task_result["importedCount"] == 1
        assert task_result["generatedLessonCount"] == 3
        task = project.list_entities("teaching_task")[0]
        assert task["course_plan_id"] == plan["id"]
        assert task["term_id"] == term["id"]
        lessons = project.list_entities("task_lesson")
        assert [item["duration_slots"] for item in lessons] == [2, 2, 1]
        assert sum(item["duration_slots"] for item in lessons) == task["weekly_slots"]
        assert project.integrity_check() == {"integrity": "ok", "foreign_key_issues": []}
    finally:
        project.close()
